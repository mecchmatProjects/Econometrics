import os
import time
import openpyxl
import numpy as np
import pandas as pd
from collections import Counter
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from matplotlib.colors import Normalize
from matplotlib.cm import ScalarMappable
from scipy.stats import pearsonr, ttest_1samp, probplot, shapiro, ttest_ind, ttest_rel, t, sem, f_oneway

INPUT_FILE = 'Data_Excel.xlsx'
PREPARED_DATA_FILE = 'Data_Excel_Bins.xlsx'


# Функція підготовки excel-таблиці до роботи
def data_preparation():
    values_to_replace = {
        "5 = Total agreement": 5,
        "DezTotal agreement=1": 1,
        "5=Excelente": 5,
        "F. Weak =1": 1,
        "F. Weak =1, 2": 1,
        "Totally disagree=1": 1,
        "5 = Total Agreement": 5,
        "5 = Totally agree": 5,
        "3, 4": 3,
        "Good": 4,
        "Neutre": 3,
        "Very Good": 5,
        "Weak": 2,
        "Very Weak": 1,
        "I can't answer": 3,
        "Agree": 4,
        "Neutru": 3,
        "Total agreement": 5,
        "DezAgree": 2,
        "Total DezAgree": 1
        # "I can't answer":3
    }

    workbook = openpyxl.load_workbook(INPUT_FILE)

    for worksheet in workbook:
        for column in worksheet.iter_cols():
            for cell in column:
                # Перевірка, чи містить клітинка текст, який потрібно замінити
                if cell.value in values_to_replace:
                    cell.value = values_to_replace[cell.value]
                # Якщо в непорожньому стовпці є порожня клітинка, вписати в неї "I can't answer"
                elif cell.value is None and any(cell1.value is not None for cell1 in column):
                    # cell.value = "I can't answer"
                    cell.value = 3
    workbook.save(PREPARED_DATA_FILE)
    workbook.close()


# Функція створення гістограм по всіх стовпцях кожного аркуша таблиці
# Не запускати без попереднього запуску функції "data_preparation()"
def histogram_creation(show_plot=True):
    workbook = pd.ExcelFile(PREPARED_DATA_FILE)

    output_directory = 'histograms'
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    for sheet_name in workbook.sheet_names:
        df = workbook.parse(sheet_name)
        for column_name in df.columns:
            column_name = str(column_name)
            if not df[column_name].dropna().empty:
                plt.figure(figsize=(8, 6))
                try:
                    plt.hist(df[column_name], bins=10, color='skyblue', edgecolor='black')
                except:
                    print(column_name, df[column_name])
                    continue
                plt.title(column_name)
                plt.xlabel('Value')
                plt.ylabel('Frequency')
                plt.grid(True)

                # Формування шляху для збереження гістограми із заміною "проблемних символів"
                clean_sheet_name = str(sheet_name).replace(' ', '_').replace('.', '').replace('/', '_').replace('"', '')
                clean_column_name = str(column_name).replace(' ', '_').replace('.', '').replace('/', '_').replace('"', '')
                histogram_filename = f'{output_directory}/{clean_sheet_name}_{clean_column_name}_histogram.png'

                plt.savefig(histogram_filename)
                if show_plot:
                    plt.show()
                    # Затримка між запитами, щоб уникнути блокування
                    time.sleep(0.2)

# Функція отримання масивів даних із стовпців excel-таблиці
# Не запускати без попереднього запуску функції "data_preparation()"
def get_arrays_of_data():
    workbook = load_workbook(PREPARED_DATA_FILE)

    # Створення словника для збереження масивів з кожного аркуша
    all_arrays = {}

    # Створення словника для збереження кількості стовпців до того, як зустрінеться перший порожній, з кожного аркуша (це використовуватиметься у побудові графіків)
    all_num_of_cols_before_empty = {}

    # Створення масиву для збереження назв стопців, щоб уникнути повторів
    array_of_column_names = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        arrays = {}
        num_of_cols_before_empty = 0

        # Знаходження к-ті стовпців до першого порожнього
        for column in sheet.iter_cols(values_only=True):
            if any(cell is not None for cell in column):
                num_of_cols_before_empty += 1
            else:
                break

        all_num_of_cols_before_empty[sheet_name] = num_of_cols_before_empty

        for column in sheet.iter_cols(values_only=True):
            # Перевірка, чи стовпець не є порожнім
            if any(cell is not None for cell in column):

                column_name = column[0]
                """
                # Перевірка, чи не опрацьований вже поточний стовпець
                if column_name not in array_of_column_names:
                    array_of_column_names.append(column_name)
                    arrays[column_name] = [cell for cell in column if cell is not None and cell != column_name]
                """
                arrays[column_name] = [cell for cell in column if cell is not None and cell != column_name]

        all_arrays[sheet_name] = arrays
    """
    for sheet_name, arrays in all_arrays.items():
        print(f"Аркуш: {sheet_name}")
        for column_name, array in arrays.items():
            print(f"\tНазва стовпця: {column_name}")
            print("\tМасив:", array)

    #print(all_num_of_cols_before_empty)
    """
    result = {
        "all_arrays": all_arrays,
        "all_num_of_cols_before_empty": all_num_of_cols_before_empty
    }

    return result


# Функція побудови графіків, використовуючи масиви даних, здобуті за допомогою функції "get_arrays_of_data()"
# Не запускати без попереднього запуску функції "data_preparation()"
def construction_of_graphs(show_plots=True):

    graphs_directory = 'graphs'
    if not os.path.exists(graphs_directory):
        os.makedirs(graphs_directory)

    letters_of_columns = {
        "PROACTIVE_ORIENTATION_[Technologies_used_are_the_latest]": "A",
        "PROACTIVE_ORIENTATION_[We_anticipate_the_potential_of_new_technologies_practices]": "B",
        "PROACTIVE_ORIENTATION_[We_systematically_try_to_acquire_and_implement_new_technologies]": "C",
        "PROACTIVE_ORIENTATION_[The_research_and_development_department_is_a_leader_in_the_field]": "D",
        "PERFORMANȚELE_FINANCIARE_[Profitul_brut]": "A",
        "PERFORMANȚELE_FINANCIARE(1)_[Profitul_brut]": "B",
        "FINANCIAL_PERFORMANCE_[Gross_Profit]": "C",
        "FINANCIAL_PERFORMANCE_[Return_on_assets]": "D",
        "FINANCIAL_PERFORMANCE_[Sales]": "E",
        "FINANCIAL_PERFORMANCE_[Earnings_per_share]": "F",
        "FINANCIAL_PERFORMANCE(1)_[Earnings_per_share]": "G",
        "FINANCIAL_PERFORMANCE_[Rate_of_Profit]": "H",
        "INNOVATION_[Research_activity]": "A",
        "INNOVATION_[The_degree_of_product_novelty]": "B",
        "INNOVATION_[Using_the_latest_technologies]": "C",
        "INNOVATION_[Speed_of_development_of_new_products]": "D",
        "INNOVATION_[Share_of_new_products_within_the_range]": "E",
        "AgeGroup": 'A',
        "EcoGroup": "A",
        "Transport": 'B',
        "T1": 'A',
        "T2": 'B',
        "T3": 'C',
        "T4": 'D',
        "CarsAgeGroup": 'A'

    }

    for sheet_name, arrays in get_arrays_of_data()["all_arrays"].items():

        # Побудова графіків для кожної пари стовпців, де х-координати - це значення зі стовпців справа, а у-координати - значення зі стовпців зліва кожного аркуша.
        counter1 = 0
        for left_column_name, left_array in arrays.items():
            # left_column_name.sorted()
            counter1 += 1
            counter2 = 0
            if counter1 <= get_arrays_of_data()["all_num_of_cols_before_empty"][sheet_name]:
                for right_column_name, right_array in arrays.items():
                    counter2 += 1
                    if counter2 > get_arrays_of_data()["all_num_of_cols_before_empty"][sheet_name]:

                        plt.figure(figsize=(15, 10))
                        point_counts = Counter(zip(right_array, left_array))

                        # Створення мапера кольорів на основі кількості повторень для кожної точки
                        max_count = max(point_counts.values())
                        norm = Normalize(vmin=0, vmax=max_count)
                        cmap = plt.get_cmap('viridis')
                        sm = ScalarMappable(norm=norm, cmap=cmap)

                        for (x, y), count in point_counts.items():
                            color = sm.to_rgba(count)
                            size = count * 65
                            plt.scatter(x, y, color=color, s=size, alpha=1)
                            plt.annotate(f'{count}', (x, y), fontsize=10, textcoords="offset points", xytext=(15, 15), ha='left')

                        #plt.scatter(right_array, left_array, color='red', alpha=1)
                        plt.title('Plot of dependancy')
                        plt.xlabel(right_column_name)
                        plt.ylabel(left_column_name)
                        plt.grid(True)

                        # Формування шляху для збереження графіка із заміною "проблемних символів"
                        clean_sheet_name = str(sheet_name).replace(' ', '_').replace('.', '').replace('/', '_').replace(
                            '"', '')
                        clean_left_column_name = str(left_column_name).replace(' ', '_').replace('.', '').replace('/',
                                                                                                        '_').replace(
                            '"', '')
                        clean_right_column_name = str(right_column_name).replace(' ', '_').replace('.', '').replace('/',
                                                                                                                  '_').replace(
                            '"', '')
                        print(f'{clean_sheet_name}_{letters_of_columns}{clean_left_column_name}_{clean_right_column_name}')
                        graph_filename = f'{graphs_directory}/{clean_sheet_name[0:2]}_{letters_of_columns[clean_left_column_name]}_{clean_right_column_name}_graph.png'

                        plt.savefig(graph_filename)
                        if show_plots:
                            plt.show()
                            # Затримка між запитами, щоб уникнути блокування
                            time.sleep(0.1)

# Функція обчислення статистичної значущості по впливу параметрів
# Не запускати без попереднього запуску функції "data_preparation()"
def correlation_calculation():
    # Словник для заміни текстових значень з таблиці на відповідні числові
    numerical_equivalent = {
        "Total DezAgree": 1,
        "Very Weak": 1,
        "DezAgree": 2,
        "Weak": 2,
        "Neutru": 3,
        "Neutre": 3,
        "Agree": 4,
        "Good": 4,
        "Total agreement": 5,
        "Very Good": 5,
        "I can't answer": 3
    }

    letters_of_columns = {
        "PROACTIVE_ORIENTATION_[Technologies_used_are_the_latest]": "A",
        "PROACTIVE_ORIENTATION_[We_anticipate_the_potential_of_new_technologies_practices]": "B",
        "PROACTIVE_ORIENTATION_[We_systematically_try_to_acquire_and_implement_new_technologies]": "C",
        "PROACTIVE_ORIENTATION_[The_research_and_development_department_is_a_leader_in_the_field]": "D",
        "PERFORMANȚELE_FINANCIARE_[Profitul_brut]": "A",
        "PERFORMANȚELE_FINANCIARE(1)_[Profitul_brut]": "B",
        "FINANCIAL_PERFORMANCE_[Gross_Profit]": "C",
        "FINANCIAL_PERFORMANCE_[Return_on_assets]": "D",
        "FINANCIAL_PERFORMANCE_[Sales]": "E",
        "FINANCIAL_PERFORMANCE_[Earnings_per_share]": "F",
        "FINANCIAL_PERFORMANCE(1)_[Earnings_per_share]": "G",
        "FINANCIAL_PERFORMANCE_[Rate_of_Profit]": "H",
        "INNOVATION_[Research_activity]": "A",
        "INNOVATION_[The_degree_of_product_novelty]": "B",
        "INNOVATION_[Using_the_latest_technologies]": "C",
        "INNOVATION_[Speed_of_development_of_new_products]": "D",
        "INNOVATION_[Share_of_new_products_within_the_range]": "E",
        "AgeGroup": 'A',
        "EcoGroup": "A",
        "Transport": 'B',
        "T1": 'A',
        "T2": 'B',
        "T3": 'C',
        "T4": 'D',
        "CarsAgeGroup": 'A'
    }

    standard_means_for_t_stat = {
        "H1": 0,
        "H2": 0,
        "H3": 0,
        "H4": 0,
        "H5": 0,
        "H6": 0,
        "H7": 0,
        "H8": 0
    }
    mean_cmp ={

        "H2": "greater",
        "H3": "greater",
    }

    with open('correlation_calculation_results.txt', 'w+', encoding='utf-8') as file:

        for sheet_name, arrays in get_arrays_of_data()["all_arrays"].items():
            counter1 = 0
            for left_column_name, left_array in arrays.items():
                counter1 += 1
                counter2 = 0
                if counter1 <= get_arrays_of_data()["all_num_of_cols_before_empty"][sheet_name]:
                    avg_diff = 0
                    avg_counter = 0
                    for right_column_name, right_array in arrays.items():
                        counter2 += 1
                        if counter2 > get_arrays_of_data()["all_num_of_cols_before_empty"][sheet_name]:
                            # Заміна текстових параметрів з лівих стовпців кожного аркуша на відповідні числові
                            numerical_left_array = []
                            for cell in left_array:
                                if cell in numerical_equivalent.keys():
                                    numerical_cell = numerical_equivalent[cell]
                                    numerical_left_array.append(numerical_cell)
                                else:
                                    numerical_left_array.append(cell)

                            # Формування заголовку для результатів із заміною "проблемних символів"
                            clean_sheet_name = str(sheet_name).replace(' ', '_').replace('.', '').replace('/',
                                                                                                          '_').replace(
                                '"', '')
                            clean_left_column_name = str(left_column_name).replace(' ', '_').replace('.',
                                                                                                     '').replace(
                                '/', '_').replace(
                                '"', '')
                            clean_right_column_name = str(right_column_name).replace(' ', '_').replace('.',
                                                                                                       '').replace(
                                '/',
                                '_').replace(
                                '"', '')
                            results_header = f'{clean_sheet_name[0:2]}_{clean_left_column_name}_{clean_right_column_name}'

                            # Обчислення кореляції Пірсона для кожної пари стовпців зліва та справа кожного аркуша.
                            correlation_coefficient, p_value = pearsonr(right_array, numerical_left_array)
                            print("\n  Results:", results_header, '\n')
                            print(f"Avg: {np.average(right_array)}, {np.average(numerical_left_array)}")
                            print(f"Var: {np.var(right_array)}, {np.var(numerical_left_array)}")

                            print("Pearson corellation:", correlation_coefficient)
                            print("p-value:", p_value)
                            # Обчислення різниці в значенні
                            difference = (np.array(right_array) - np.array(numerical_left_array))/10

                            # Проведення t-тесту
                            # Pop_mean = 1 if results_header.startswith('H1') or results_header.startswith('H1') else 0

                            pop_mean = np.round(np.average(4*difference))/4 # standard_means_for_t_stat.get(clean_sheet_name[0:2],0)

                            t_statistic, p_value_ttest = ttest_1samp(difference, pop_mean)
                            print(f"Difference stat: {pop_mean}\n")
                            print("t-statistics:", t_statistic)
                            print("p-value for t-test:", p_value_ttest)

                            # Perform the Shapiro-Wilk test
                            stat_s, p_value_s = shapiro(difference)

                            # Output the results
                            print(f"Shapiro-Wilk Test Statistic: {stat_s}")
                            print(f"P-value: {p_value_s}")
                            alpha = 0.05  # Common significance level
                            if p_value_s > alpha:
                                print("Sample looks normally distributed (fail to reject H0)")
                            else:
                                print("Sample does not look normally distributed (reject H0)")
                            # probplot(difference,plot=plt)
                            # plt.show()
                            # Perform independent two-sample t-test
                            stat_i, p_value_i = ttest_ind(right_array, numerical_left_array,
                                                      # alternative='less',
                                                      equal_var=False)  # Use equal_var=True if variances are assumed equal

                            # Output the results
                            print(f"Independent t-test statistic: {stat_i}")
                            print(f"P-value: {p_value_i}")

                            # Perform paired t-test
                            alternate = mean_cmp.get(clean_sheet_name[0:2], "two-sided")
                            print(f"checking {alternate}")
                            stat_p, p_value_p = ttest_rel(right_array, numerical_left_array, alternative=alternate)

                            # Output the results
                            print(f"Paired t-test statistic: {stat_p}")
                            print(f"P-value: {p_value_p}")


                            # Interpretation
                            alpha = 0.05
                            if p_value_p > alpha:
                                print("No significant difference in means (fail to reject H0)")
                            else:
                                print("Significant difference in means (reject H0)")
                            print('\n')

                            # Calculate mean and standard error of the differences
                            mean_diff = np.mean(difference)
                            se_diff = sem(difference)  # Standard error of the mean of differences

                            # Degrees of freedom for paired t-test
                            df = len(difference) - 1

                            # Define confidence level
                            confidence_level = 0.95
                            t_critical = t.ppf((1 + confidence_level) / 2, df)

                            # Calculate confidence interval for the mean difference
                            ci_lower = mean_diff - t_critical * se_diff
                            ci_upper = mean_diff + t_critical * se_diff

                            print(
                                f"{int(confidence_level * 100)}% Confidence Interval for the mean difference: ({ci_lower}, {ci_upper})")

                            # Interpretation
                            if ci_lower > 0:
                                print(
                                    "The confidence interval does not include zero, indicating dist1's mean is significantly greater than dist2's mean.")
                            elif ci_upper < 0:
                                print(
                                    "The confidence interval does not include zero, indicating dist1's mean is significantly less than dist2's mean.")
                            else:
                                print(
                                    "The confidence interval includes zero, suggesting no significant difference in means.")

                            file.write(f"\n{results_header}\n")
                            file.write(f"Pearson c.c: {correlation_coefficient}\n")
                            file.write(f"p-value: {p_value}\n")

                            file.write(f"Difference stat: {pop_mean}\n")
                            file.write(f"t-stat for differences: {t_statistic}\n")
                            file.write(f"p-value for t-test: {p_value_ttest}\n\n")
                            # file.write(f"Shapiro-Wilk Test Statistic: {stat_s}")
                            # file.write(f"P-value: {p_value_s}")
                            file.write(f"Paired t-test statistic: {stat_p}\n")
                            file.write(f"P-value: {p_value_p}\n")
                            if p_value_p > alpha:
                                file.write(f"{alternate} is failed to reject H0)\n")
                            else:
                                file.write(f" {alternate} in means (reject H0)\n")

                            if ci_lower > 0:
                                file.write(
                                    "The confidence interval does not include zero, indicating dist1's mean is significantly greater than dist2's mean.\n")
                            elif ci_upper < 0:
                                file.write(
                                    "The confidence interval does not include zero, indicating dist1's mean is significantly less than dist2's mean.\n")
                            else:
                                file.write(
                                    "The confidence interval includes zero, suggesting no significant difference in means.\n")



                            from scipy.stats import rankdata, kendalltau, chi2_contingency, mannwhitneyu


                            # Step 1: Rank the data
                            rank_X = rankdata(numerical_left_array)
                            rank_Y = rankdata(right_array)

                            # Step 2: Compute the differences between ranks
                            d = rank_X - rank_Y

                            # Step 3: Square the differences
                            d_squared = d ** 2

                            # Step 4: Sum the squared differences
                            sum_d_squared = np.sum(d_squared)

                            # Step 5: Calculate Spearman's rank correlation coefficient
                            n = len(right_array)
                            r_s = 1 - (6 * sum_d_squared) / (n * (n ** 2 - 1))

                            print(f"Spearman’s Rank Correlation Coefficient: {r_s}")
                            file.write(f"Spearman’s Rank Correlation Coefficient: {r_s}\n")

                            # Calculate Kendall's Tau using scipy
                            tau, p_value = kendalltau(numerical_left_array, right_array)

                            print(f"Kendall’s Tau: {tau} {p_value}")
                            file.write(f"SKendall’s Tau, p-value: {tau}, {p_value}\n")

                            # Create a contingency table
                            contingency_table = pd.crosstab(numerical_left_array, right_array)

                            # Perform the Chi-Square test
                            chi2, p, dof, expected = chi2_contingency(contingency_table)

                            # Calculate Cramér's V
                            n = np.sum(contingency_table.values)  # Total number of observations
                            min_dimension = min(contingency_table.shape) - 1  # min(k-1, r-1)
                            cramers_v = np.sqrt(chi2 / (n * min_dimension))

                            print(f"Chi-Square Statistic: {chi2}")
                            print(f"P-Value: {p}")
                            print(f"Cramér's V: {cramers_v}")

                            file.write(f"Chi-Square Statistic, P-value, Cramér's V: {chi2}, {p}, {cramers_v}\n")

                            # Perform the Mann - Whitney U-Test

                            stat_mu, p_value_mu = mannwhitneyu(numerical_left_array, right_array, alternative=alternate)

                            # Output the results
                            print(f"Mann-Whitney U Test statistic: {stat_mu}")
                            print(f"P-value: {p_value_mu}")

                            # Interpretation
                            alpha = 0.05
                            if p_value_mu < alpha:
                                print(
                                    "There is a significant difference between the distributions of dist1 and dist2 (reject H0)")
                            else:
                                print(
                                    "No significant difference between the distributions of dist1 and dist2 (fail to reject H0)")

                            if clean_sheet_name[0:2]=="H5" and clean_left_column_name.startswith("T"):
                                index_notransp = np.array(np.where(np.array(numerical_left_array) <= 1.5)[0]).astype(int)
                                index_transp = np.array(np.where(np.array(numerical_left_array) >= 1.5)[0]).astype(int)
                                # print(indexes1)
                                data1 = np.array(right_array)[index_notransp]
                                data2 = np.array(right_array)[index_transp]
                                print(f"{np.average(data1)}, {data1.shape}, {np.average(data2)}, {data2.shape}")
                                file.write(f"Average Transport vs NonTransport:{np.average(data1)}, {np.average(data2)}\n")

                                alternate2 = 'greater'
                                stat_mu, p_value_mu = mannwhitneyu(data1, data2,
                                                                   alternative=alternate2)

                                # Output the results
                                print(f"Mann-Whitney U Test statistic: {stat_mu}")
                                print(f"P-value: {p_value_mu}")
                                file.write(f"Mann-Whitney U Test statistic: {stat_mu}  ")
                                file.write(f"P-value: {p_value_mu}\n")

                                # Interpretation
                                alpha = 0.05
                                if p_value_mu < alpha:
                                    file.write("reject H0\n")
                                else:
                                    file.write("fail to reject H0\n")

                                # Calculate Effect Size (r = Z / sqrt(N))
                                N = len(data1) + len(data2)  # Total sample size
                                effect_size = stat_mu / np.sqrt(N)  # Effect size calculation

                                # Output the Effect Size
                                print(f"Effect Size (r): {effect_size}")
                                file.write(f"Effect Size (r): {effect_size}\n")

                                # Interpretation of Effect Size
                                if effect_size < 0.1:
                                    print("Negligible effect")
                                    file.write("Negligible effect\n")
                                elif 0.1 <= effect_size < 0.3:
                                    print("Small effect")
                                    file.write("Small effect\n")
                                elif 0.3 <= effect_size < 0.5:
                                    print("Medium effect")
                                    file.write("Medium effect\n")
                                else:
                                    print("Large effect")
                                    file.write("Large effect\n")

                                # Perform One-Way ANOVA
                                anova_stat, anova_p_value = f_oneway(data1, data2)

                                # Output the results of One-Way ANOVA
                                print(f"One-Way ANOVA F-statistic: {anova_stat}")
                                print(f"P-value: {anova_p_value}")
                                file.write(f"One-Way ANOVA F-statistic: {anova_stat}  ")
                                file.write(f"P-value: {anova_p_value}\n")

                                # Interpretation of ANOVA Results
                                if anova_p_value < alpha:
                                    print("Significant difference between group means (reject H0)")
                                    file.write("Significant difference between group means (reject H0)\n")
                                else:
                                    print("No significant difference between group means (fail to reject H0)")
                                    file.write("No significant difference between group means (fail to reject H0)\n")





# data_preparation()
# histogram_creation(show_plot=False)
# construction_of_graphs(show_plots=False)
correlation_calculation()
